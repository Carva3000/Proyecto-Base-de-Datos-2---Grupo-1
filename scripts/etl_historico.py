import psycopg2
from psycopg2.extras import execute_values
import openpyxl
from datetime import date, datetime
import logging
import sys
import sys
from pathlib import Path
sys.path.insert(0, str(Path(__file__).parent.parent / 'config'))
from config import DB_CONFIG, LOG_FILE

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(LOG_FILE, encoding='utf-8'),
        logging.StreamHandler(sys.stdout)
    ]
)
logger = logging.getLogger(__name__)


def get_connection():
    return psycopg2.connect(**DB_CONFIG)


def get_or_create_dimension(cursor, table, key_column, value_column, value):
    """Get existing dimension key or create new one"""
    cursor.execute(f"SELECT {key_column} FROM {table} WHERE {value_column} = %s", (value,))
    result = cursor.fetchone()
    if result:
        return result[0]
    cursor.execute(
        f"INSERT INTO {table} ({value_column}) VALUES (%s) RETURNING {key_column}",
        (value,)
    )
    return cursor.fetchone()[0]


def get_or_create_prenda(cursor, nombre, color, talla, categoria):
    """Get or create prenda dimension"""
    cursor.execute(
        """SELECT Id_Prenda FROM Dim_Prenda 
           WHERE Nombre_Prenda = %s AND Color = %s AND Talla = %s AND Categoria = %s""",
        (nombre, color, talla, categoria)
    )
    result = cursor.fetchone()
    if result:
        return result[0]
    cursor.execute(
        """INSERT INTO Dim_Prenda (Nombre_Prenda, Color, Talla, Categoria) 
           VALUES (%s, %s, %s, %s) RETURNING Id_Prenda""",
        (nombre, color, talla, categoria)
    )
    return cursor.fetchone()[0]


def get_or_create_tiempo(cursor, fecha):
    """Get or create tiempo dimension"""
    if isinstance(fecha, str):
        fecha = datetime.strptime(fecha, '%Y-%m-%d').date()
    elif isinstance(fecha, datetime):
        fecha = fecha.date()
    
    cursor.execute("SELECT Id_Tiempo FROM Dim_Tiempo WHERE Fecha = %s", (fecha,))
    result = cursor.fetchone()
    if result:
        return result[0]
    
    dia = fecha.day
    mes = fecha.month
    anio = fecha.year
    trimestre = (mes - 1) // 3 + 1
    dia_semana = fecha.isoweekday()
    es_fin_semana = dia_semana in (6, 7)
    
    cursor.execute(
        """INSERT INTO Dim_Tiempo (Fecha, Dia, Mes, Anio, Trimestre, Dia_Semana, Es_Fin_Semana)
           VALUES (%s, %s, %s, %s, %s, %s, %s) RETURNING Id_Tiempo""",
        (fecha, dia, mes, anio, trimestre, dia_semana, es_fin_semana)
    )
    return cursor.fetchone()[0]


def read_excel_data(filepath, sheet_name=None):
    """Read Excel file and return rows as list of dicts"""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active
    
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is not None:
            rows.append(dict(zip(headers, row)))
    return rows


def process_row(row):
    """Process a single row and return structured data"""
    fecha = row['Fecha_Venta']
    if isinstance(fecha, str):
        fecha = datetime.strptime(fecha, '%Y-%m-%d').date()
    elif isinstance(fecha, datetime):
        fecha = fecha.date()
    
    precio = float(row['Precio_Unitario']) if row['Precio_Unitario'] else 0
    cantidad = int(row['Cantidad']) if row['Cantidad'] else 0
    descuento = float(row['Descuento']) if row['Descuento'] else 0
    monto_linea = precio * cantidad * (1 - descuento / 100)
    
    return {
        'fecha': fecha,
        'vendedor': row['Vendedor'],
        'metodo_pago': row['Metodo_Pago'],
        'promocion': row['Nombre_Promo'],
        'prenda': row['Nombre_Prenda'],
        'color': row['Color'],
        'talla': row['Talla'],
        'categoria': row['Categoria'],
        'cantidad': cantidad,
        'precio_unitario': precio,
        'descuento': descuento,
        'monto_linea': monto_linea
    }


def load_historical_data(excel_path, sheet_name='Historico_14_Meses'):
    """Main function to load historical data"""
    logger.info(f"Iniciando carga histórica desde {excel_path}")
    
    rows = read_excel_data(excel_path, sheet_name)
    logger.info(f"Leídas {len(rows)} filas del Excel")
    
    conn = get_connection()
    conn.autocommit = False
    cursor = conn.cursor()
    
    try:
        dim_prenda_cache = {}
        dim_vendedor_cache = {}
        dim_formapago_cache = {}
        dim_promocion_cache = {}
        dim_tiempo_cache = {}
        
        fact_ventas = {}
        fact_ingresos = {}
        fact_transacciones = {}
        fact_promos = {}
        fact_rendimiento = {}
        
        for i, row in enumerate(rows, 1):
            data = process_row(row)
            
            # Dimensiones con cache
            if data['vendedor'] not in dim_vendedor_cache:
                dim_vendedor_cache[data['vendedor']] = get_or_create_dimension(
                    cursor, 'Dim_Vendedor', 'Id_Vendedor', 'Nombre_Vendedor', data['vendedor']
                )
            id_vendedor = dim_vendedor_cache[data['vendedor']]
            
            if data['metodo_pago'] not in dim_formapago_cache:
                dim_formapago_cache[data['metodo_pago']] = get_or_create_dimension(
                    cursor, 'Dim_FormaPago', 'Id_FormaPago', 'Tipo_FormaPago', data['metodo_pago']
                )
            id_formapago = dim_formapago_cache[data['metodo_pago']]
            
            if data['promocion'] not in dim_promocion_cache:
                dim_promocion_cache[data['promocion']] = get_or_create_dimension(
                    cursor, 'Dim_Promocion', 'Id_Promocion', 'Nombre_Promocion', data['promocion']
                )
            id_promocion = dim_promocion_cache[data['promocion']]
            
            prenda_key = (data['prenda'], data['color'], data['talla'], data['categoria'])
            if prenda_key not in dim_prenda_cache:
                dim_prenda_cache[prenda_key] = get_or_create_prenda(
                    cursor, data['prenda'], data['color'], data['talla'], data['categoria']
                )
            id_prenda = dim_prenda_cache[prenda_key]
            
            if data['fecha'] not in dim_tiempo_cache:
                dim_tiempo_cache[data['fecha']] = get_or_create_tiempo(cursor, data['fecha'])
            id_tiempo = dim_tiempo_cache[data['fecha']]
            
            # Acumular hechos
            # Fact_Ventas_Prendas
            key_vp = (id_prenda, id_tiempo)
            if key_vp not in fact_ventas:
                fact_ventas[key_vp] = {'cantidad': 0, 'monto': 0}
            fact_ventas[key_vp]['cantidad'] += data['cantidad']
            fact_ventas[key_vp]['monto'] += data['monto_linea']
            
            # Fact_Ingresos_Temporales
            if id_tiempo not in fact_ingresos:
                fact_ingresos[id_tiempo] = 0
            fact_ingresos[id_tiempo] += data['monto_linea']
            
            # Fact_Transacciones (contar facturas únicas por día y método de pago)
            key_trans = (id_formapago, id_tiempo)
            if key_trans not in fact_transacciones:
                fact_transacciones[key_trans] = set()
            fact_transacciones[key_trans].add(row['Nro_Factura'])
            
            # Fact_Venta_Promociones
            key_promo = (id_prenda, id_promocion, id_tiempo)
            if key_promo not in fact_promos:
                fact_promos[key_promo] = {'cantidad': 0, 'monto': 0}
            fact_promos[key_promo]['cantidad'] += data['cantidad']
            fact_promos[key_promo]['monto'] += data['monto_linea']
            
            # Hechos_Rendimiento
            key_rend = (id_vendedor, id_tiempo)
            if key_rend not in fact_rendimiento:
                fact_rendimiento[key_rend] = 0
            fact_rendimiento[key_rend] += data['monto_linea']
            
            if i % 100 == 0:
                logger.info(f"Procesadas {i}/{len(rows)} filas")
        
        logger.info("Insertando datos en tablas de hechos...")
        
        # Insertar Fact_Ventas_Prendas
        valores_vp = [(k[0], k[1], v['cantidad'], round(v['monto'], 2)) 
                      for k, v in fact_ventas.items()]
        execute_values(cursor, 
            """INSERT INTO Fact_Ventas_Prendas (Id_Prenda, Id_Tiempo, Cantidad_Vendida, Monto_Vendido)
               VALUES %s
               ON CONFLICT (Id_Prenda, Id_Tiempo) DO UPDATE SET
               Cantidad_Vendida = EXCLUDED.Cantidad_Vendida,
               Monto_Vendido = EXCLUDED.Monto_Vendido""",
            valores_vp)
        
        # Insertar Fact_Ingresos_Temporales
        valores_ing = [(k, round(v, 2)) for k, v in fact_ingresos.items()]
        execute_values(cursor,
            """INSERT INTO Fact_Ingresos_Temporales (Id_Tiempo, Monto_Total_Venta)
               VALUES %s
               ON CONFLICT (Id_Tiempo) DO UPDATE SET
               Monto_Total_Venta = EXCLUDED.Monto_Total_Venta""",
            valores_ing)
        
        # Insertar Fact_Transacciones
        valores_trans = [(k[0], k[1], len(v)) for k, v in fact_transacciones.items()]
        execute_values(cursor,
            """INSERT INTO Fact_Transacciones (Id_FormaPago, Id_Tiempo, Cantidad_Transacciones)
               VALUES %s
               ON CONFLICT (Id_FormaPago, Id_Tiempo) DO UPDATE SET
               Cantidad_Transacciones = EXCLUDED.Cantidad_Transacciones""",
            valores_trans)
        
        # Insertar Fact_Venta_Promociones
        valores_promo = [(k[0], k[1], k[2], v['cantidad'], round(v['monto'], 2)) 
                         for k, v in fact_promos.items()]
        execute_values(cursor,
            """INSERT INTO Fact_Venta_Promociones (Id_Prenda, Id_Promocion, Id_Tiempo, Cantidad_Vendida, Monto_Vendido)
               VALUES %s
               ON CONFLICT (Id_Prenda, Id_Promocion, Id_Tiempo) DO UPDATE SET
               Cantidad_Vendida = EXCLUDED.Cantidad_Vendida,
               Monto_Vendido = EXCLUDED.Monto_Vendido""",
            valores_promo)
        
        # Insertar Hechos_Rendimiento
        valores_rend = [(k[0], k[1], round(v, 2)) for k, v in fact_rendimiento.items()]
        execute_values(cursor,
            """INSERT INTO Hechos_Rendimiento (Id_Vendedor, Id_Tiempo, Monto_Total_Venta)
               VALUES %s
               ON CONFLICT (Id_Vendedor, Id_Tiempo) DO UPDATE SET
               Monto_Total_Venta = EXCLUDED.Monto_Total_Venta""",
            valores_rend)
        
        conn.commit()
        logger.info("Carga histórica completada exitosamente")
        
    except Exception as e:
        conn.rollback()
        logger.error(f"Error en carga histórica: {e}")
        raise
    finally:
        cursor.close()
        conn.close()


if __name__ == '__main__':
    from config import EXCEL_HISTORICO
    load_historical_data(EXCEL_HISTORICO)